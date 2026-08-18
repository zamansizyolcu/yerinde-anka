"""
actions/olcek_hazirla.py — Ders içi katılım ölçeği / proje değerlendirme
ölçeği (dönem/sene sonu teslim edilen) hazırlama.

Bu araç, öğretmenin kullandığı hazır bir Excel değerlendirme aracını
(50 sayfalık, tamamı FORMÜLLERLE birbirine bağlı — VBA makrosu YOK,
doğrulandı) yeni bir sınıf/dönem için klonlar. Mekanizma:

  - "Anasayfa" sekmesindeki 9 sabit bilgi hücresi (eğitim yılı, dönem,
    okul/idareci/öğretmen bilgisi, ders adı, sınıf adı) TÜM diğer
    sayfalardaki başlıkları ve imza alanlarını CONCATENATE formülleriyle
    besliyor — bu yüzden sadece bu hücreleri güncellemek yeterli, 50
    sayfaya tek tek dokunmaya gerek yok.
  - "Eokul" sayfası, öğretmenin e-okul sisteminden kopyalayıp yapıştırdığı
    GERÇEK not verisinin girildiği yer (Okul No + Adı Soyadı + sınav/proje
    notları). Bu notlar Claude'un bilemeyeceği gerçek veriler olduğundan,
    bu araç sadece İSTENİRSE öğrenci No/Ad-Soyad listesini önceden doldurur
    (notlar boş bırakılır — öğretmen e-okul'dan yapıştırarak tamamlar,
    tıpkı şu ana kadar yaptığı gibi).

Bilişim ve Robotik için KRİTERLER AYNI (aynı şablon, tek fark Anasayfa'daki
"DERSİN ADI" hücresi) — bu yüzden ayrı bir şablon/mantık gerekmiyor.
"""

from __future__ import annotations

import platform
import re
import subprocess
import time
from pathlib import Path

from app_config import get_app_config_value

_IS_WINDOWS = platform.system() == "Windows"

DERS_ADLARI = {
    "bilişim": "BİLİŞİM TEKNOLOJİLERİ VE YAZILIM",
    "robotik": "SEÇMELİ ROBOTİK KODLAMA",
}
OGRETMEN_UNVANI = "BİLİŞİM TEKNOLOJİLERİ ÖĞRETMENİ"

# Anasayfa'daki sabit bilgi hücreleri (referans dosyada doğrulandı)
_ANASAYFA_HUCRELERI = {
    "egitim_yili": "E5", "donem": "H6", "okul_adi": "E7",
    "idareci_adi": "E8", "idareci_unvani": "E9", "ders_adi": "E10",
    "ogretmen_adi": "E11", "ogretmen_unvani": "E12", "sinif_adi": "E13",
}


def _tr_lower(s: str) -> str:
    return s.replace("İ", "i").replace("I", "ı").lower()


def _tr_upper(s: str) -> str:
    """Python'un genel .upper() metodu Türkçe i/ı çiftinde hata yapar (ör.
    'Yeğitoğlu'.upper() yanlışlıkla 'YEĞITOĞLU' üretir, 'YEĞİTOĞLU' değil)."""
    return s.replace("i", "İ").replace("ı", "I").upper()


def _normalize_ders(ders: str) -> str:
    d = _tr_lower(ders or "")
    return "robotik" if "robot" in d else "bilişim"


def _open_file(path: Path) -> None:
    if _IS_WINDOWS:
        import os
        os.startfile(str(path))  # noqa: S606
    else:
        subprocess.Popen(["xdg-open", str(path)],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def _top_left(ws, row: int, col: int):
    for rng in ws.merged_cells.ranges:
        if (row, col) in rng.cells:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return ws.cell(row=row, column=col)


def _load_workbook_saglam(path: Path):
    """Bazı gerçek dünya .xlsx dosyaları (ör. bozuk birleştirilmiş hücre
    referansı) openpyxl ile doğrudan açılamayabiliyor. Böyle durumda
    LibreOffice ile sessizce yeniden kaydedip (bu, dosyayı onarır) tekrar
    dener."""
    import openpyxl
    try:
        return openpyxl.load_workbook(str(path))
    except Exception:
        pass
    import tempfile
    tmp_dir = Path(tempfile.mkdtemp(prefix="yerinde_olcek_"))
    skill_soffice = Path("/mnt/skills/public/docx/scripts/office/soffice.py")
    soffice_cmd = ["soffice"]
    if skill_soffice.exists():
        import sys
        soffice_cmd = [sys.executable, str(skill_soffice)]
    subprocess.run(soffice_cmd + ["--headless", "--convert-to", "xlsx",
                                    "--outdir", str(tmp_dir), str(path)],
                   capture_output=True, timeout=120)
    onarilmis = tmp_dir / path.name
    if not onarilmis.exists():
        raise ValueError("Dosya açılamadı ve otomatik onarım da başarısız oldu.")
    return openpyxl.load_workbook(str(onarilmis))


_OGRENCI_SATIR_RE = re.compile(r"^\s*(\d+)?\s*[.\)\-]?\s*(.+?)\s*$")


def _puantaj_sheet_bul(wb, sinif: str, ders_norm: str) -> str | None:
    """Puantaj dosyasındaki 8 sınıf sayfası arasından ('6A BİLİŞİM',
    '5C BİLİŞM' gibi yazım farklılıkları/eksik harfler dahil) istenen
    sınıf+dersi bulur."""
    hedef_sinif = re.sub(r"[^0-9A-ZÇĞİÖŞÜ]", "", _tr_upper(sinif))  # '5/A' -> '5A'
    hedef_anahtar = "ROBOT" if ders_norm == "robotik" else "BİLİŞ"
    for name in wb.sheetnames:
        norm = _tr_upper(name)
        norm_sinif_kismi = re.sub(r"[^0-9A-ZÇĞİÖŞÜ]", "", norm)
        if hedef_sinif and hedef_sinif in norm_sinif_kismi and hedef_anahtar in norm:
            return name
    return None


def _puantaj_ogrencileri_oku(wb, sheet_adi: str) -> list[dict]:
    ws = wb[sheet_adi]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    baslik = [_tr_lower(str(c or "")).strip() for c in rows[0]]

    def _sutun(*anahtarlar):
        for i, h in enumerate(baslik):
            if any(a in h for a in anahtarlar):
                return i
        return None

    c_no = _sutun("okul no", "no")
    c_ad = _sutun("adı soyadı", "ad soyad")
    c_s1 = _sutun("1. sınav", "1.sınav")
    c_s2 = _sutun("2. sınav", "2.sınav")
    c_p1 = _sutun("1. proje", "1.proje")

    sonuc = []
    for row in rows[1:]:
        if c_ad is None or c_ad >= len(row) or not row[c_ad]:
            continue
        sonuc.append({
            "no": row[c_no] if c_no is not None and c_no < len(row) else None,
            "ad": str(row[c_ad]).strip(),
            "s1": row[c_s1] if c_s1 is not None and c_s1 < len(row) else None,
            "s2": row[c_s2] if c_s2 is not None and c_s2 < len(row) else None,
            "p1": row[c_p1] if c_p1 is not None and c_p1 < len(row) else None,
        })
    return sonuc


def _ogrencileri_ayristir(metin: str) -> list[tuple[str, str]]:
    """'17 Nur Cennet Zangalı' / '24) Beyzanur Patlıcan' gibi satırları
    (okul_no, ad_soyad) çiftlerine ayırır. Numara verilmemişse okul_no boş
    kalır."""
    sonuc = []
    for satir in re.split(r"[\n;]+", metin):
        satir = satir.strip().strip(",")
        if not satir:
            continue
        m = re.match(r"^(\d+)\s*[.\)\-]?\s*(.+)$", satir)
        if m:
            sonuc.append((m.group(1), _tr_upper(m.group(2).strip())))
        else:
            sonuc.append(("", _tr_upper(satir)))
    return sonuc


def olcek_hazirla(
    ders: str = "bilişim",
    sinif: str = "5/A",
    donem: str = "",
    egitim_yili: str = "",
    ogrenciler: str = "",
    puantaj_dosya_yolu: str = "",
    dosya_yolu: str = "",
) -> str:
    from actions.belge_referanslari import referans_yolu

    ders_norm = _normalize_ders(ders)
    ref_path = (Path(dosya_yolu) if dosya_yolu and Path(dosya_yolu).exists()
                else referans_yolu("olcek_sablonu"))
    if not ref_path:
        return ("Ders içi katılım / proje değerlendirme ölçeği örneğini bulamadım. "
                "Önce 'DOSYA YÜKLE' ile örnek .xlsx'i yükleyip 'bunu ölçek şablonu "
                "olarak kaydet' de, sonra tekrar iste.")

    try:
        wb = _load_workbook_saglam(ref_path)
    except Exception as e:
        return f"Şablon dosyası okunamadı: {e}"

    if "Anasayfa" not in wb.sheetnames:
        return "Bu dosyada beklenen 'Anasayfa' sekmesini bulamadım — gerçek şablon mu?"
    ws = wb["Anasayfa"]

    if not egitim_yili:
        from datetime import datetime
        y = datetime.now().year
        egitim_yili = f"{y}-{y+1}" if datetime.now().month >= 7 else f"{y-1}-{y}"
    if not donem:
        donem = "1.DÖNEM"
    else:
        donem = _tr_upper(donem.strip())
        if donem.isdigit():
            donem = f"{donem}.DÖNEM"

    # Mevcut Anasayfa değerlerini oku (okul/idareci/öğretmen bilgisi gibi
    # değişmeyecek alanları KORUMAK için — sadece ders/sınıf/dönem/yıl
    # güncellenir, geri kalan referans dosyadan aynen devam eder).
    ws[_ANASAYFA_HUCRELERI["egitim_yili"]] = egitim_yili
    ws[_ANASAYFA_HUCRELERI["donem"]] = donem
    ws[_ANASAYFA_HUCRELERI["ders_adi"]] = DERS_ADLARI[ders_norm]
    ws[_ANASAYFA_HUCRELERI["sinif_adi"]] = _tr_upper(sinif.strip())
    # Öğretmen ünvanı, ders ne olursa olsun aynı resmî unvan (Emre Dumancı
    # örneğinde olduğu gibi tek bir branş ünvanı altında iki dersi de veriyor)
    mevcut_unvan = str(ws[_ANASAYFA_HUCRELERI["ogretmen_unvani"]].value or "").strip()
    if not mevcut_unvan:
        ws[_ANASAYFA_HUCRELERI["ogretmen_unvani"]] = OGRETMEN_UNVANI

    ogrenci_sayisi = 0
    puantaj_kullanildi = False
    if "Eokul" in wb.sheetnames:
        eokul = wb["Eokul"]

        puantaj_path = (Path(puantaj_dosya_yolu) if puantaj_dosya_yolu and Path(puantaj_dosya_yolu).exists()
                         else referans_yolu("puantaj"))
        satirlar = None
        if puantaj_path:
            try:
                puantaj_wb = _load_workbook_saglam(puantaj_path)
                sheet_adi = _puantaj_sheet_bul(puantaj_wb, sinif, ders_norm)
                if sheet_adi:
                    ogrenciler_ham = _puantaj_ogrencileri_oku(puantaj_wb, sheet_adi)
                    if ogrenciler_ham:
                        satirlar = ogrenciler_ham
                        puantaj_kullanildi = True
            except Exception:
                satirlar = None

        if satirlar:
            for i, ogr in enumerate(satirlar[:30]):
                satir = 5 + i * 2
                if ogr["no"] is not None:
                    _top_left(eokul, satir, 2).value = ogr["no"]
                _top_left(eokul, satir, 3).value = _tr_upper(ogr["ad"])
                if ogr["s1"] is not None:
                    _top_left(eokul, satir, 4).value = ogr["s1"]
                if ogr["s2"] is not None:
                    _top_left(eokul, satir, 5).value = ogr["s2"]
                if ogr["p1"] is not None:
                    _top_left(eokul, satir, 6).value = ogr["p1"]
                # D/E/F sütunları öğrenci başına İKİ satırı KAPSAMIYOR (sadece
                # No/Ad Soyadı birleşik) — referans şablonda alt satırda
                # (satir+1) kalıntı örnek veriler bulunabiliyor (doğrulandı).
                # Karışıklık olmasın diye alt satırı temizliyoruz.
                for col in (4, 5, 6):
                    try:
                        _top_left(eokul, satir + 1, col).value = None
                    except Exception:
                        pass
            ogrenci_sayisi = len(satirlar[:30])
        elif ogrenciler.strip():
            cifler = _ogrencileri_ayristir(ogrenciler)
            for i, (no, ad) in enumerate(cifler[:30]):  # şablon en fazla 30 öğrenci destekliyor
                # Her öğrenci İKİ satırlık birleşik hücreleri kapsıyor (B5:B6,
                # B7:B8, B9:B10, ...) — adım 2 olmalı, yoksa art arda yazılan
                # değerler aynı birleşik hücrenin üzerine düşüp birbirini
                # ezer (doğrulandı: gerçek dosyada test edip bulduk).
                satir = 5 + i * 2
                if no:
                    hucre = _top_left(eokul, satir, 2)
                    hucre.value = int(no) if no.isdigit() else no
                hucre2 = _top_left(eokul, satir, 3)
                hucre2.value = ad
            ogrenci_sayisi = len(cifler[:30])

    from actions.code_tools import ensure_workspace_folder
    folder = ensure_workspace_folder("Değerlendirme Ölçekleri")
    stamp = time.strftime("%Y-%m-%d %H.%M")
    sinif_dosya_adi = re.sub(r"[\\/:*?\"<>|]", "-", sinif.strip())
    out_name = f"{ders_norm.capitalize()} {sinif_dosya_adi} Ders İçi ve Proje Ölçeği {stamp}.xlsx"
    out_path = folder / out_name
    i = 2
    while out_path.exists():
        out_path = folder / f"{ders_norm.capitalize()} {sinif_dosya_adi} Ders İçi ve Proje Ölçeği {stamp} ({i}).xlsx"
        i += 1

    try:
        wb.save(str(out_path))
    except Exception as e:
        return f"Dosya kaydedilemedi: {e}"

    try:
        _open_file(out_path)
    except Exception:
        pass

    if puantaj_kullanildi:
        ogrenci_notu = (f" Puantaj dosyandan {ogrenci_sayisi} öğrencinin adı, okul no'su ve "
                         f"mevcut sınav/proje notları otomatik dolduruldu. NOT: 'Ders İçi "
                         f"Katılım' puanları puantajında tek bir özet sayı olarak duruyor "
                         f"(ör. '100'), ama ölçek şablonundaki Ders İçi 1/2/3 sayfaları her "
                         f"öğrenciyi 5 ayrı kazanıma göre 1-5 arası puanlıyor — bu yüzden o "
                         f"kısmı senin ilgili sayfalarda elle işaretlemen gerekiyor, uydurma "
                         f"bir dağılım yapmadım.")
    elif ogrenci_sayisi:
        ogrenci_notu = (f" {ogrenci_sayisi} öğrenci adı/no önceden dolduruldu — "
                         f"notları her zamanki gibi e-okul'dan kopyalayıp Eokul "
                         f"sayfasındaki kırmızı hücreye (B5) yapıştırman yeterli.")
    else:
        ogrenci_notu = (" Öğrenci listesi boş bırakıldı — e-okul'dan kopyalayıp "
                         "Eokul sayfasına her zamanki gibi yapıştır.")

    return (f"'{DERS_ADLARI[ders_norm]}' dersi, {sinif} şubesi, {egitim_yili} "
            f"{donem} için ders içi katılım ve proje değerlendirme ölçeği "
            f"hazırlandı ve açıldı: {out_path.name}.{ogrenci_notu}")
