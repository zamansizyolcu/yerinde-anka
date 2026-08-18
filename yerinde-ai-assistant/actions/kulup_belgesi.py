"""
actions/kulup_belgesi.py — Öğrenci kulübü yıllık çalışma planı üretimi.

MEB Sosyal Etkinlikler Yönetmeliği EK-7/b ("ÖĞRENCİ KULÜBÜ SOSYAL ETKİNLİKLER
YILLIK ÇALIŞMA PLANI") formatına uygun, okulun kendi doldurduğu örnek Excel
dosyasını (AY -> o ayki ETKİNLİKLER + BELİRLİ GÜN VE HAFTALAR) yeni eğitim-
öğretim yılı için klonlar: yılı, katılımcı öğrenci sayılarını ve (istenirse)
etkinlik içeriklerini günceller — tablo yapısı, kenarlıklar, birleştirilmiş
hücreler AYNEN korunur (openpyxl ile var olan hücreler üzerinde değişiklik
yapılır, sıfırdan üretilmez).

NOT — Kulüp TÜZÜĞÜ ayrı bir belge: MEB'in genel sosyal etkinlik yönetmeliği
metnidir, kulübe/döneme özel değildir, yıldan yıla değişmez — bu yüzden bu
modül tüzüğü DEĞİL, sadece yıllık çalışma planını üretir. Faaliyet Raporu
(dönem/yıl SONUNDA, 'yapıldı' diliyle özet) ayrı, gelecekte eklenecek bir
işlev — bu modülün ürettiği plan, o rapor için de temel oluşturacak.
"""

from __future__ import annotations

import json
import platform
import re
import subprocess
import time
from pathlib import Path

import requests

from app_config import get_app_config_value, get_model_provider, ollama_think_value

_IS_WINDOWS = platform.system() == "Windows"
LLM_TIMEOUT = 600
_AY_ADLARI = ["EYLÜL", "EKİM", "KASIM", "ARALIK", "OCAK", "ŞUBAT", "MART",
              "NİSAN", "MAYIS", "HAZİRAN", "TEMMUZ", "AĞUSTOS"]


def _tr_lower(s: str) -> str:
    return s.replace("İ", "i").replace("I", "ı").lower()


def _tr_upper_first(word: str) -> str:
    if not word:
        return word
    first = word[0]
    if first == "i":
        first_upper = "İ"
    elif first == "ı":
        first_upper = "I"
    else:
        first_upper = first.upper()
    return first_upper + _tr_lower(word[1:])


def _tr_title(s: str) -> str:
    """Python'un .title() metodu Türkçe İ/I çiftinde bozuk sonuç üretebiliyor
    (ör. 'BİLİŞİM'.title() -> yanlışlıkla birleşik nokta karakteri ekliyor).
    Kelime kelime, doğru Türkçe eşlemeyle başa çıkıyoruz."""
    return " ".join(_tr_upper_first(w) for w in s.split(" ") if w) if s else s


def _open_file(path: Path) -> None:
    if _IS_WINDOWS:
        import os
        os.startfile(str(path))  # noqa: S606
    else:
        subprocess.Popen(["xdg-open", str(path)],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def _top_left(ws, row: int, col: int):
    """Bir hücre birleştirilmiş bir aralığın parçasıysa, o aralığın sol-üst
    hücresini döner (openpyxl'de sadece oraya yazılabilir)."""
    for rng in ws.merged_cells.ranges:
        if (row, col) in rng.cells:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return ws.cell(row=row, column=col)


# ══ LLM yardımcıları (zümre/sınav modülleriyle aynı desen) ═══════════════
def _extract_json_obj(raw: str) -> dict | None:
    raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip())
    m = re.search(r"\{.*\}", raw, re.S)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except Exception:
        return None


def _call_gemini_text(prompt: str) -> str | None:
    api_key = str(get_app_config_value("gemini_api_key", "") or "").strip()
    if not api_key:
        return None
    try:
        from google import genai
    except ImportError:
        return None
    try:
        client = genai.Client(api_key=api_key)
        for model in ("models/gemini-2.5-flash", "models/gemini-2.0-flash", "models/gemini-2.5-flash-lite"):
            try:
                resp = client.models.generate_content(model=model, contents=[prompt])
                if resp and resp.text:
                    return resp.text
            except Exception:
                continue
    except Exception:
        pass
    return None


def _call_ollama_text(prompt: str, num_predict: int = 1600) -> str | None:
    try:
        host = str(get_app_config_value("ollama_host", "http://localhost:11434") or "http://localhost:11434")
        model = str(get_app_config_value("ollama_model", "llama3.1") or "llama3.1")
        payload = {"model": model, "stream": False,
                   "messages": [{"role": "user", "content": prompt}],
                   "options": {"temperature": 0.6, "num_predict": num_predict},
                   "think": ollama_think_value(), "keep_alive": "30m"}
        r = requests.post(f"{host}/api/chat", json=payload, timeout=LLM_TIMEOUT)
        if r.status_code >= 400:
            payload.pop("think", None)
            r = requests.post(f"{host}/api/chat", json=payload, timeout=LLM_TIMEOUT)
        return r.json().get("message", {}).get("content", "")
    except Exception:
        return None


def _call_llm(prompt: str, num_predict: int = 1600) -> str | None:
    return (_call_gemini_text(prompt) if get_model_provider() == "gemini"
            else _call_ollama_text(prompt, num_predict))


def _etkinlikleri_yenile(mevcut_aylar: dict, ctx: dict) -> dict | None:
    aylar_metni = "\n".join(f"- {ay}: {icerik}" for ay, icerik in mevcut_aylar.items())
    ek = f"\nEK TALİMAT: {ctx['ek_talimat']}" if ctx.get("ek_talimat") else ""
    prompt = (
        f"Sen bir Türk devlet ortaokulunda '{ctx['kulup_adi']}' öğrenci "
        f"kulübünün danışman öğretmenisin. Aşağıda GEÇEN YILKİ aylık etkinlik "
        f"planı var. {ctx['egitim_yili']} eğitim-öğretim yılı için, AYNI "
        f"ay yapısını koruyarak, benzer ruhta ama TAZE/güncellenmiş etkinlik "
        f"fikirleriyle yeni bir plan yaz (geçen yılın birebir kopyası olmasın, "
        f"ama aynı pedagojik akışı — Eylül'de kulüp kurulumu, Haziran'da "
        f"değerlendirme gibi — koru).{ek}\n\n"
        f"GEÇEN YILKİ PLAN:\n{aylar_metni}\n\n"
        f"SADECE şu JSON nesnesini döndür, başka HİÇBİR şey yazma: "
        f'{{"EYLÜL": "madde1\\nmadde2\\n...", "EKİM": "...", ...}} — '
        f"geçen yılki planda hangi aylar varsa TAM OLARAK aynı ay isimleri "
        f"anahtar olarak kullanılsın, değerler her biri 2-4 maddelik "
        f"kısa satırlar olsun (\\n ile ayrılmış)."
    )
    raw = _call_llm(prompt, num_predict=1800)
    obj = _extract_json_obj(raw) if raw else None
    if not obj:
        raw2 = _call_llm(prompt + "\n\nUNUTMA: SADECE JSON nesnesi.", num_predict=1800) if raw else None
        obj = _extract_json_obj(raw2) if raw2 else None
    return obj


# ══ Ana giriş noktası ═══════════════════════════════════════════════════
def kulup_calisma_plani_olustur(
    egitim_yili: str = "",
    katilimci_toplam: str = "",
    katilimci_kiz: str = "",
    katilimci_erkek: str = "",
    danisman_adi: str = "",
    etkinlikleri_yenile: bool = False,
    ek_talimat: str = "",
    dosya_yolu: str = "",
) -> str:
    from actions.belge_referanslari import referans_yolu

    ref_path = (Path(dosya_yolu) if dosya_yolu and Path(dosya_yolu).exists()
                else referans_yolu("kulup_yillik_calisma_plani"))
    if not ref_path:
        return ("Örnek bir kulüp yıllık çalışma planı (.xlsx) bulamadım. Önce "
                "'DOSYA YÜKLE' ile geçen senenin planını yükleyip 'bunu kulüp "
                "yıllık çalışma planı olarak kaydet' de, sonra tekrar iste.")

    try:
        import openpyxl
    except ImportError:
        return "Excel dosyalarıyla çalışmak için 'pip install openpyxl' gerekli."

    try:
        wb = openpyxl.load_workbook(str(ref_path))
    except Exception as e:
        return f"Örnek dosya okunamadı: {e}"

    ws = wb[wb.sheetnames[0]]

    # Bugünün tarihine göre makul bir varsayılan eğitim yılı
    if not egitim_yili:
        from datetime import datetime
        y = datetime.now().year
        egitim_yili = f"{y}-{y+1}" if datetime.now().month >= 7 else f"{y-1}-{y}"

    # Başlık hücresini bul (A1 civarı, "EĞİTİM" ve "ÖĞRETİM" geçen ilk hücre)
    baslik_cell = None
    kulup_adi = ""
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            if cell.value and "EĞİTİM" in str(cell.value).upper() and "ÖĞRETİM" in str(cell.value).upper():
                baslik_cell = cell
                break
        if baslik_cell:
            break
    if baslik_cell:
        eski_metin = str(baslik_cell.value)
        yeni_metin = re.sub(r"\b(\d{4})\s*[-–]\s*(\d{4})\b", egitim_yili, eski_metin)
        tl = _top_left(ws, baslik_cell.row, baslik_cell.column)
        tl.value = yeni_metin
        m = re.search(r"(?:ORTAOKULU|İLKOKULU|LİSESİ|ANAOKULU)\s+(.+?KULÜBÜ)", eski_metin.upper())
        if m:
            kulup_adi = _tr_title(m.group(1).strip())
    if not kulup_adi:
        kulup_adi = "Öğrenci"

    # Katılımcı sayıları — gerçek metin "KATILIMCI ÖĞRENCİ SAYISI" (toplam için
    # ayrı bir anahtar kelime YOK), "KATILIMCI KIZ ..." ve "KATILIMCI ERKEK ..."
    # şeklinde geliyor; toplamı KIZ/ERKEK içermeyen satır olarak ayırt ediyoruz.
    for row in ws.iter_rows(min_row=1, max_row=6):
        for cell in row:
            if not cell.value or "KATILIMCI" not in str(cell.value).upper():
                continue
            metin_upper = str(cell.value).upper()
            if "KIZ" in metin_upper:
                yeni_sayi = katilimci_kiz
            elif "ERKEK" in metin_upper:
                yeni_sayi = katilimci_erkek
            else:
                yeni_sayi = katilimci_toplam
            if yeni_sayi:
                tl = _top_left(ws, cell.row, cell.column)
                etiket = re.split(r":\s*\d*\s*$", str(cell.value))[0].rstrip()
                tl.value = f"{etiket}: {yeni_sayi}"

    # Danışman öğretmen adı
    if danisman_adi:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and _tr_lower(str(cell.value)).strip() == "danışman öğretmen":
                    alt_hucre = ws.cell(row=cell.row + 1, column=cell.column)
                    tl = _top_left(ws, alt_hucre.row, alt_hucre.column)
                    tl.value = danisman_adi.upper()

    # AY satırlarını bul ve (istenirse) etkinlikleri yenile
    ay_satirlari = {}  # AY_ADI -> (row, col_etkinlik, mevcut_metin)
    etkinlik_col = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value or "").strip().upper()
            if val in _AY_ADLARI:
                # etkinlik sütunu genelde bir sonraki sütun
                ec = cell.column + 1
                ecell = ws.cell(row=cell.row, column=ec)
                ay_satirlari[val] = (cell.row, ec, str(ecell.value or ""))
                etkinlik_col = ec

    ctx = {"kulup_adi": kulup_adi, "egitim_yili": egitim_yili, "ek_talimat": ek_talimat.strip()}
    yenileme_basarisiz = False
    if etkinlikleri_yenile and ay_satirlari:
        mevcut = {ay: bilgi[2] for ay, bilgi in ay_satirlari.items()}
        yeni = _etkinlikleri_yenile(mevcut, ctx)
        if yeni:
            for ay, (row, col, _eski) in ay_satirlari.items():
                if ay in yeni and yeni[ay]:
                    tl = _top_left(ws, row, col)
                    tl.value = str(yeni[ay]).replace("\\n", "\n")
                    # Kaynak dosyada bazı ay hücrelerinde satır kaydırma kapalı
                    # olabiliyor (tek satırlık eski içerik yeterliydi) — yeni
                    # çok satırlı içerik doğru görünsün diye zorluyoruz.
                    from copy import copy
                    yeni_align = copy(tl.alignment)
                    yeni_align.wrap_text = True
                    tl.alignment = yeni_align
        else:
            yenileme_basarisiz = True

    from actions.code_tools import ensure_workspace_folder
    folder = ensure_workspace_folder("Kulüp Belgeleri")
    stamp = time.strftime("%Y-%m-%d %H.%M")
    out_name = f"{kulup_adi} Yıllık Çalışma Planı {egitim_yili} {stamp}.xlsx"
    out_path = folder / out_name
    i = 2
    while out_path.exists():
        out_path = folder / f"{kulup_adi} Yıllık Çalışma Planı {egitim_yili} {stamp} ({i}).xlsx"
        i += 1

    try:
        wb.save(str(out_path))
    except Exception as e:
        return f"Dosya kaydedilemedi: {e}"

    try:
        _open_file(out_path)
    except Exception:
        pass

    uyari = ""
    if etkinlikleri_yenile and yenileme_basarisiz:
        uyari = " NOT: Etkinlik içerikleri yenilenmeye çalışıldı ama model yanıtı ayrıştırılamadı — geçen yılkiler aynen kaldı, istersen elle güncelle."

    return (f"'{kulup_adi}' {egitim_yili} yıllık çalışma planı hazırlandı "
            f"ve açıldı: {out_path.name}.{uyari}")
